import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os, sys

application_path = os.path.dirname(sys.executable)
month = input('Introduce month: ')

# сводная таблица Считает общую сумму продукции на гендер
# df = pd.read_csv('supermarket_sales - Sheet1.csv')
# df = df[['Gender','Product line','Total']]
# pivot_table = df.pivot_table(index='Gender', columns='Product line',
#                              values='Total', aggfunc='sum').round(0)
# pivot_table.to_excel('pivot_table.xlsx','Report',startrow=4)

# создание графиков из данных таблицы
input_path = os.path.join(application_path, 'pivot_table.xlsx')
wb = load_workbook(input_path)
sheet = wb['Report'] # Лист Excel
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row
barchart = BarChart()
data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart, "B12")
barchart.title = "Sales by Product line"
barchart.style = 5
wb.save('barchart.xlsx')

# создание формул Excel
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'
wb.save('report.xlsx')

# создание заголовков в Excel
wb = load_workbook('report.xlsx')
sheet = wb['Report']
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

output_path = os.path.join(application_path, f'report_{month}.xlsx')
wb.save(output_path)